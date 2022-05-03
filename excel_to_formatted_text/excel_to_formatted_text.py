import argparse
import pathlib
import os
import hashlib
from openpyxl import load_workbook
from jinja2 import Template


def main():

    parser = argparse.ArgumentParser(description='Excel to formated text')
    parser.add_argument('-x', '--excelfile', required=True, help='Excel filename')
    parser.add_argument('-s', '--sheetname', required=True, help='Sheet name')
    parser.add_argument('-o', '--outputfile', required=True, help='Output filename format')
    parser.add_argument('-oe', '--outputfile-encoding', default='utf8', required=False, help='Output text file encoding')
    parser.add_argument('-lt', '--outputfile-lineterminator', choices=['cr', 'lf', 'crlf'],
                        default='lf', required=False, help='Output text file line terminator')
    parser.add_argument('-t', '--templatefile', required=True, help='Template text filename')
    parser.add_argument('-te', '--templatefile-encoding', default='utf8', required=False, help='Template text file encoding')
    parser.add_argument('-r', '--startrow', default=1, type=int, required=False, help='Excel sheet start row no')
    parser.add_argument('-c', '--startcol', default=1, type=int, required=False, help='Excel sheet start column no')
    parser.add_argument('-b', '--blank-skip-columns', required=False, type=str, action='append',
                        help='Skip if this column is blank')

    args = parser.parse_args()

    print(f'Excel file : {args.excelfile}')
    print(f'Sheet name : {args.sheetname}')
    print(f'Output file format : {args.outputfile}')
    print(f'Output file encoding : {args.outputfile_encoding}')
    print(f'Output file lineterminator : {args.outputfile_lineterminator}')
    print(f'Template file : {args.templatefile}')
    print(f'Template file encoding : {args.templatefile_encoding}')
    print(f'Start row : {args.startrow}')
    print(f'Start column : {args.startcol}')
    print(f'Blank skip column : {args.blank_skip_columns}')

    lineterminator = '\n'
    if args.outputfile_lineterminator == 'cr':
        lineterminator = '\r'
    elif args.outputfile_lineterminator == 'crlf':
        lineterminator = '\r\n'

    to_text(args.excelfile, args.sheetname, args.startrow, args.startcol, args.blank_skip_columns,
            args.outputfile, args.outputfile_encoding, lineterminator,
            args.templatefile, args.templatefile_encoding)


def to_text(excelfile, sheetname, startrow, startcol, blank_skip_columns,
            filename_format, outputfile_encoding, lineterminator,
            templatefile, templatefile_encoding):
    wb = load_workbook(filename=excelfile, read_only=True)
    ws = wb[sheetname]

    col_names = []
    for col in range(startcol, ws.max_column+1):
        col_names.append((col-1, ws.cell(startrow, col).value))
        print(ws.cell(startrow, col).value)

    with open(templatefile, 'rt', encoding=templatefile_encoding) as f:
        template_text = f.read()
    template = Template(template_text, newline_sequence=lineterminator, keep_trailing_newline=True)

    for row in ws.iter_rows(min_row=startrow+1):
        cell_values = to_dictionary(row, col_names)
        if not is_outputtable_row(cell_values, blank_skip_columns):
            continue

        filename = generate_output_filename(cell_values, filename_format)
#        print(filename)

        text = template.render(cell_values)

        if is_modified(text, filename):
            print(f'{filename} is modified.')
            with open(filename, 'wt', newline='', encoding=outputfile_encoding) as f:
                f.write(text)
        else:
            print(f'{filename} is not modified.')


def is_outputtable_row(cell_values, blank_skip_columns):
    if blank_skip_columns is None:
        return True
    for col_name in blank_skip_columns:
        if col_name in cell_values \
                and (cell_values[col_name] is None or cell_values[col_name] == ''):
            return False
    return True


def generate_output_filename(cell_values, filename_format):
    output_filename = filename_format
    for col_name in cell_values:
        output_filename = output_filename.replace('{' + col_name + '}', cell_values[col_name])

    output_filename = pathlib.Path(output_filename).resolve()
    os.makedirs(os.path.dirname(output_filename), exist_ok=True)

    return pathlib.Path(output_filename).resolve()


def to_dictionary(row, col_names):
    dict = {}
    for col_no, col_name in col_names:
        if not row[col_no].value is None:
            dict[col_name] = str(row[col_no].value)
        else:
            dict[col_name] = ''

    return dict


def is_modified(new_text, filename):
    if not os.path.exists(filename):
        return True

    with open(filename, 'rb') as file:
        file_text = file.read()
        file_text_hs = hashlib.sha256(file_text).hexdigest()

    # ファイルのエンコードが UTF-8 と仮定、new_text は UTF-8 のはず
    new_text_hs = hashlib.sha256(new_text.encode()).hexdigest()
    if new_text_hs == file_text_hs:
        return False

    # ファイルのエンコードが SJIS と仮定、new_text は UTF-8 のはずなので sjis に変換
    new_text_hs = hashlib.sha256(new_text.encode('sjis')).hexdigest()
    return not (new_text_hs == file_text_hs)


if __name__ == "__main__":
    main()
